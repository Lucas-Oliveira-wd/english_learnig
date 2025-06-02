from flask import Flask, request, jsonify
from openpyxl import load_workbook, Workbook
from flask_cors import CORS
import os

app = Flask(__name__)
CORS(app)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))  # /servidor
ARQUIVO = os.path.join(BASE_DIR, "../banco_de_dados/vocabulario.xlsx")

def carregar_dados():
    if not os.path.exists(ARQUIVO):
        os.makedirs(os.path.dirname(ARQUIVO), exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.append(["termo", "classe", "traducao", "acertos"])
        wb.save(ARQUIVO)

    wb = load_workbook(ARQUIVO)
    ws = wb.active
    dados = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        dados.append({
            "termo": row[0],
            "classe": row[1],
            "traducao": row[2],
            "acertos": row[3]
        })
    return dados

def salvar_dados(dados):
    wb = Workbook()
    ws = wb.active
    ws.append(["termo", "classe", "traducao", "acertos"])
    for item in dados:
        ws.append([
            item["termo"], item["classe"], item["traducao"], item["acertos"]
        ])
    wb.save(ARQUIVO)

@app.route("/dados", methods=["GET"])
def get_dados():
    return jsonify(carregar_dados())

@app.route("/dados", methods=["POST"])
def post_dado():
    novo = request.json
    dados = carregar_dados()

    if any(p["termo"].lower() == novo["termo"].lower() for p in dados):
        return jsonify({"mensagem": f'A palavra "{novo["termo"]}" já foi cadastrada!'}), 200

    dados.append(novo)
    salvar_dados(dados)
    return jsonify({"status": "ok"})

@app.route("/dados", methods=["PUT"])
def atualizar_dados():
    novos_dados = request.json
    salvar_dados(novos_dados)
    return jsonify({"status": "atualizado"})

if __name__ == "__main__":
    app.run(debug=True)
